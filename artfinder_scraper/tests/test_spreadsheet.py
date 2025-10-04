"""Unit coverage for the spreadsheet export helpers."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PIL import Image as PILImage

from artfinder_scraper.scraping.models import Artwork
from artfinder_scraper.scraping.spreadsheet import (
    DEFAULT_SHEET_NAME,
    SPREADSHEET_COLUMNS,
    append_artwork_to_spreadsheet,
)


def _create_artwork(
    title: str,
    source_url: str,
    *,
    image_path: Path | None = None,
    description: str | None = None,
) -> Artwork:
    return Artwork(
        title=title,
        description=description if description is not None else f"Description for {title}",
        price_gbp="£75",
        size="40 x 40 cm",
        sold=False,
        image_url=None,
        image_path=str(image_path) if image_path else None,
        materials_used="Oil on canvas",
        source_url=source_url,
    )


def _make_image(path: Path) -> Path:
    image = PILImage.new("RGB", (640, 480), color=(12, 34, 56))
    image.save(path)
    return path


def test_append_artwork_creates_workbook_and_embeds_image(tmp_path: Path) -> None:
    workbook_path = tmp_path / "artworks.xlsx"
    image_path = _make_image(tmp_path / "example.png")
    artwork = _create_artwork(
        "Sunrise",
        "https://example.com/product/sunrise/",
        image_path=image_path,
    )

    appended = append_artwork_to_spreadsheet(artwork, workbook_path)
    assert appended is True

    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    assert worksheet.title == DEFAULT_SHEET_NAME
    header_values = [
        worksheet.cell(row=1, column=index + 1).value
        for index in range(len(SPREADSHEET_COLUMNS))
    ]
    assert header_values == list(SPREADSHEET_COLUMNS)

    row_values = [
        worksheet.cell(row=2, column=index + 1).value
        for index in range(len(SPREADSHEET_COLUMNS))
    ]
    assert row_values[0] is None
    assert row_values[1] == artwork.title
    assert row_values[2] == artwork.size
    assert row_values[3] == artwork.medium
    assert row_values[4] == artwork.materials_used
    assert row_values[5] == "£75"
    assert row_values[6] == artwork.description
    assert row_values[7] == "for sale"
    assert row_values[8] == str(artwork.source_url)

    hyperlink = worksheet.cell(row=2, column=len(SPREADSHEET_COLUMNS)).hyperlink
    assert hyperlink is not None
    assert hyperlink.target == str(artwork.source_url)

    images = getattr(worksheet, "_images", [])
    assert len(images) == 1
    anchor = images[0].anchor
    assert getattr(anchor._from, "row") == 1
    assert getattr(anchor._from, "col") == 0

    workbook.close()


def test_append_artwork_skips_duplicates(tmp_path: Path) -> None:
    workbook_path = tmp_path / "artworks.xlsx"

    first = _create_artwork("First", "https://example.com/product/first/")
    duplicate_slug = _create_artwork("Second", "https://example.com/product/first/")
    duplicate_title = _create_artwork("First", "https://example.com/product/second/")
    second = _create_artwork("Second", "https://example.com/product/second/")

    assert append_artwork_to_spreadsheet(first, workbook_path) is True
    assert append_artwork_to_spreadsheet(duplicate_slug, workbook_path) is False
    assert append_artwork_to_spreadsheet(duplicate_title, workbook_path) is False
    assert append_artwork_to_spreadsheet(second, workbook_path) is True

    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    assert worksheet.max_row == 3  # header + two unique entries
    titles = [worksheet.cell(row=row, column=2).value for row in range(2, worksheet.max_row + 1)]
    assert titles == ["First", "Second"]
    workbook.close()


def test_description_line_breaks_are_preserved(tmp_path: Path) -> None:
    workbook_path = tmp_path / "artworks.xlsx"
    multiline_description = "First paragraph.\r\n\r\nSecond paragraph with detail."
    artwork = _create_artwork(
        "Layered",
        "https://example.com/product/layered/",
        description=multiline_description,
    )

    assert append_artwork_to_spreadsheet(artwork, workbook_path) is True

    workbook = load_workbook(workbook_path)
    worksheet = workbook.active

    cell = worksheet.cell(row=2, column=7)
    assert cell.value == multiline_description.replace("\r\n", "\n")
    assert cell.alignment.wrapText is True

    workbook.close()
