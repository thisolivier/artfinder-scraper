"""Compose spreadsheet outputs summarizing scraped artworks and metadata."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from decimal import Decimal
from typing import Sequence
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

from .models import Artwork

DEFAULT_SHEET_NAME = "Artworks"
DEFAULT_MAX_IMAGE_DIMENSION = 320
SPREADSHEET_COLUMNS: Sequence[str] = (
    "image",
    "title",
    "size",
    "medium",
    "materials used",
    "price",
    "description",
    "status",
    "art finder link",
)


@dataclass(frozen=True)
class SpreadsheetState:
    """Capture the existing identifiers in a worksheet for deduplication."""

    titles: frozenset[str]
    slugs: frozenset[str]


def append_artwork_to_spreadsheet(
    artwork: Artwork,
    workbook_path: Path,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
    max_image_dimension: int = DEFAULT_MAX_IMAGE_DIMENSION,
) -> bool:
    """Append ``artwork`` to ``workbook_path`` and return ``True`` when added.

    The workbook is created when missing, headers are enforced each call, and
    duplicate rows are skipped whenever an existing slug or title is detected.
    Images are resized to fit within ``max_image_dimension`` pixels in either
    direction before being embedded in the first column.
    """

    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = _load_or_create_workbook(workbook_path)
    worksheet = _get_or_create_worksheet(workbook, sheet_name)
    _ensure_headers(worksheet)

    state = _collect_existing_identifiers(worksheet)
    slug_key = artwork.slug.casefold()
    title_key = artwork.title.casefold()
    if slug_key in state.slugs or title_key in state.titles:
        _close_workbook(workbook)
        return False

    worksheet.append(_build_row_payload(artwork))
    row_index = worksheet.max_row
    _apply_hyperlink(worksheet, row_index, str(artwork.source_url))
    _preserve_description_line_breaks(worksheet, row_index)

    if artwork.image_path:
        excel_image = _load_image_for_excel(Path(artwork.image_path), max_image_dimension)
        if excel_image is not None:
            anchor = f"A{row_index}"
            worksheet.add_image(excel_image, anchor)
            worksheet.row_dimensions[row_index].height = max(
                worksheet.row_dimensions[row_index].height or 0,
                excel_image.height * 0.75,
            )

    workbook.save(workbook_path)
    _close_workbook(workbook)
    return True


def _load_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = DEFAULT_SHEET_NAME
    _ensure_headers(worksheet)
    return workbook


def _get_or_create_worksheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    worksheet = workbook.create_sheet(title=sheet_name)
    if len(workbook.sheetnames) > 1 and "Sheet" in workbook.sheetnames:
        default_sheet = workbook["Sheet"]
        if default_sheet.max_row == 1 and default_sheet.max_column == 1 and not default_sheet["A1"].value:
            workbook.remove(default_sheet)
    return worksheet


def _ensure_headers(worksheet: Worksheet) -> None:
    header_values = [worksheet.cell(row=1, column=index + 1).value for index in range(len(SPREADSHEET_COLUMNS))]
    if header_values != list(SPREADSHEET_COLUMNS):
        for column_index, header in enumerate(SPREADSHEET_COLUMNS, start=1):
            worksheet.cell(row=1, column=column_index, value=header)

    if worksheet.max_column > len(SPREADSHEET_COLUMNS):
        worksheet.delete_cols(len(SPREADSHEET_COLUMNS) + 1, worksheet.max_column - len(SPREADSHEET_COLUMNS))

    worksheet.column_dimensions["A"].width = 36
    worksheet.freeze_panes = "B2"


def _collect_existing_identifiers(worksheet: Worksheet) -> SpreadsheetState:
    titles: set[str] = set()
    slugs: set[str] = set()

    for row in worksheet.iter_rows(min_row=2, max_col=len(SPREADSHEET_COLUMNS), values_only=True):
        title = row[1] if len(row) > 1 else None
        link = row[8] if len(row) > 8 else None

        if isinstance(title, str) and title:
            titles.add(title.casefold())

        slug = _slug_from_url(link) if isinstance(link, str) else None
        if slug:
            slugs.add(slug.casefold())

    return SpreadsheetState(titles=frozenset(titles), slugs=frozenset(slugs))


def _build_row_payload(artwork: Artwork) -> Sequence[str | None]:
    return (
        None,
        artwork.title,
        artwork.size,
        artwork.medium,
        artwork.materials_used,
        _format_price(artwork.price_gbp),
        artwork.description,
        "sold" if artwork.sold else "for sale",
        str(artwork.source_url),
    )


def _format_price(value: Decimal | str | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        return cleaned if cleaned.startswith("£") else f"£{cleaned}"
    return f"£{value}"


def _apply_hyperlink(worksheet: Worksheet, row_index: int, url: str) -> None:
    cell = worksheet.cell(row=row_index, column=len(SPREADSHEET_COLUMNS))
    cell.value = url
    cell.hyperlink = url
    cell.style = "Hyperlink"


def _preserve_description_line_breaks(worksheet: Worksheet, row_index: int) -> None:
    column_index = SPREADSHEET_COLUMNS.index("description") + 1
    cell = worksheet.cell(row=row_index, column=column_index)

    if isinstance(cell.value, str):
        normalized_value = cell.value.replace("\r\n", "\n")
        if normalized_value != cell.value:
            cell.value = normalized_value

        if "\n" in cell.value:
            cell.alignment = Alignment(wrap_text=True)


def _load_image_for_excel(image_path: Path, max_dimension: int) -> ExcelImage | None:
    if not image_path.exists() or max_dimension <= 0:
        return None

    try:
        with PILImage.open(image_path) as source_image:
            image = source_image.copy()
    except OSError:
        return None

    image.thumbnail((max_dimension, max_dimension))

    buffer = BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)

    excel_image = ExcelImage(buffer)
    excel_image.width, excel_image.height = image.size
    excel_image._buffer = buffer  # type: ignore[attr-defined]
    return excel_image


def _close_workbook(workbook: Workbook) -> None:
    close_callable = getattr(workbook, "close", None)
    if callable(close_callable):
        close_callable()


def _slug_from_url(url: str | None) -> str | None:
    if not url:
        return None

    parsed = urlparse(url)
    segments = [segment for segment in parsed.path.split("/") if segment]
    if not segments:
        return None

    if "product" in segments:
        try:
            index = segments.index("product")
            if index + 1 < len(segments):
                return segments[index + 1]
        except ValueError:
            pass
    return segments[-1]


__all__ = ["append_artwork_to_spreadsheet", "SPREADSHEET_COLUMNS", "DEFAULT_SHEET_NAME"]
